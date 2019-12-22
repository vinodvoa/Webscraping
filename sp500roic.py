
import requests
import time
import json
import logging
import openpyxl as pyxl

from bs4 import BeautifulSoup

################################################################################
# Logging setup (hierarchy: DIWEC)
################################################################################
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

formatter = logging.Formatter('%(levelname)s:%(name)s:%(asctime)s:%(message)s')

stream_handler = logging.StreamHandler()
stream_handler.setLevel(logging.INFO)
stream_handler.setFormatter(formatter)
logger.addHandler(stream_handler)

file_handler = logging.FileHandler('stocksdebug.log')
file_handler.setFormatter(formatter)
file_handler.setLevel(logging.DEBUG)
logger.addHandler(file_handler)

file_handler = logging.FileHandler('stockserr.log')
file_handler.setFormatter(formatter)
file_handler.setLevel(logging.ERROR)
logger.addHandler(file_handler)

SOURCE = '/Users/vinodverghese/Dropbox/Python/Learning/sp500.xlsx'

logger.info('Loading workbook')

try:
    wb = pyxl.load_workbook(SOURCE)

except Exception as e:
    logger.exception('Workbook load error : %s' % e)
    quit()

logger.info('Loading complete')

################################################################################
# Get US stock prices
################################################################################
logger.info('Getting ROIC of US S&P500 stocks..')

ws = wb['ROIC']

row = 1

while (ws.cell(row, 1).value):
    usticker = ws.cell(row, 1).value
    logger.info('US stock ticker : %s' % usticker)

    URL = 'http://www.stock2own.com/StockAnalysis/Stock/US/' + usticker + '?w = 1349'
    # logger.debug('US stock price url : %s' % URL)

    page = requests.get(URL)

    soup = BeautifulSoup(page.content, 'lxml')

    lis = soup.findAll('li')

    for lis in lis:
        if 'ROIC' in str(lis.contents):
            ws.cell(row, 3).value = lis.span.text
            logger.info('ROIC : %s' % ws.cell(row, 3).value)

    # logger.debug('Row : %s' % row)
    row += 1
    time.sleep(2)

###############################################################################
# save workbook
###############################################################################
logger.info('Saving workbook..')

try:
    wb.save(SOURCE)

except Exception as e:
    logger.exception('Workbook save error : %s' % e)
    quit()

logger.info('Saved')
