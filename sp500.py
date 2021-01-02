#!/Users/vinodverghese/anaconda3/bin/ python3

import requests
import time
import openpyxl as pyxl

from bs4 import BeautifulSoup

SOURCE = '/Volumes/Secomba/vinodverghese/Boxcryptor/Dropbox/Personal/Finance/sp500.xlsx'
WIKIURL = 'https://en.wikipedia.org/wiki/List_of_S%26P_500_companies'
SRCURL = 'https://finance.yahoo.com/quote/'
SP500SHEET = 'sp500'
DELAY = 2

wb = None
ws = None


def load_fs_workbook():
    """ Load workbook """
    print('Loading fs workbook')

    try:
        wb = pyxl.load_workbook(filename='sp500.xlsx')

    except Exception as e:
        print('Workbook load error : %s' % e)
        quit()

    print('Loading of fs complete')

    return wb


def build_sp500_list():
    # get SP500 list from WIKI
    page = requests.get(WIKIURL)

    soup = BeautifulSoup(page.content)
    # print(soup.body)

    table = soup.find('table', {'class': 'wikitable sortable'})
    # print(table)

    row = 2

    for trow in table.find_all('tr')[1:]:
        ticker = trow.find_all('td')[0].text.rstrip()
        company = trow.find_all('td')[1].text.rstrip()
        hq = trow.find_all('td')[5].text.rstrip()
        added = trow.find_all('td')[6].text.rstrip()
        founded = trow.find_all('td')[8].text.rstrip()

        print(ticker, company, hq, added, founded)

        ws['A' + str(row)] = ticker
        ws['B' + str(row)] = company
        ws['C' + str(row)] = hq
        ws['D' + str(row)] = added
        ws['E' + str(row)] = founded

        row += 1

    time.sleep(2)


def get_other_info():
    row = 2

    while (ws.cell(row, 1).value) and (ws.cell(row, 6).value is None):
        print(row, ws.cell(row, 1).value, ws.cell(row, 6).value)

        # Data from Yahoo
        yahoourl = SRCURL + ws.cell(row, 1).value + '?p=' + ws.cell(row, 1).value
        # print(yahoourl)

        page = requests.get(yahoourl)

        soup = BeautifulSoup(page.content, 'lxml')

        # # Price
        # try:
        #     divtag = soup.find('div', class_='My(6px) Pos(r) smartphone_Mt(6px)')

        # except Exception as e:
        #     print('Find error : %s' % e)
        #     row += 1
        #     continue

        # try:
        #     price = divtag.div.span.text

        # except Exception as e:
        #     print('Tag not found : %s' % e)
        #     row += 1
        #     continue

        # ws['F' + str(row)] = float(price.replace(',', ''))
        # # print('%s / %s' % (ticker, ws['F' + str(row)])

        # Fair Value
        try:
            valuediv = soup.findAll('div', class_='Fw(b) Fl(end)--m Fz(s) C($primaryColor')
            ws['O' + str(row)] = valuediv[0].text

        except Exception as e:
            print('Fair Value find error : %s' % e)
            row += 1
            continue

        # Pattern
        try:
            patdiv = soup.find('div', class_='W(1/4)--mobp W(1/2) IbBox')

        except Exception as e:
            print('Pattern find error : %s' % e)
            row += 1
            continue

        try:
            ws['P' + str(row)] = patdiv.div.span.text

        except Exception as e:
            print('Pattern index error : %s' % e)
            row += 1
            continue

        time.sleep(1)

        # Stats
        yahoourl = SRCURL + ws.cell(row, 1).value + '/key-statistics?p=' + ws.cell(row, 1).value
        # print('URL : %s' % yahoourl)

        page = requests.get(yahoourl)

        soup = BeautifulSoup(page.content, 'lxml')

        try:
            tds = soup.findAll('td', class_='Ta(c) Pstart(10px) Miw(60px) Miw(80px)--pnclg Bgc($lv1BgColor) fi-row:h_Bgc($hoverBgColor)')

        except Exception as e:
            print('Stats find error : %s' % e)
            row += 1
            continue

        try:
            ws['F' + str(row)] = tds[0].text
            ws['G' + str(row)] = tds[1].text
            ws['H' + str(row)] = tds[2].text
            ws['I' + str(row)] = tds[3].text
            ws['J' + str(row)] = tds[4].text
            ws['K' + str(row)] = tds[5].text
            ws['L' + str(row)] = tds[6].text
            ws['M' + str(row)] = tds[7].text
            ws['N' + str(row)] = tds[8].text

        except Exception as e:
            print('Stats index error : %s' % e)
            row += 1
            continue

        time.sleep(1)

        # Sector/Industry
        yahoourl = SRCURL + ws.cell(row, 1).value + '/profile?p=' + ws.cell(row, 1).value
        # print('URL : %s' % yahoourl)

        page = requests.get(yahoourl)

        soup = BeautifulSoup(page.content, 'lxml')

        try:
            tds = soup.findAll('span', class_='Fw(600)')

        except Exception as e:
            print('Sector find error : %s' % e)
            row += 1
            continue

        try:
            ws['Q' + str(row)] = tds[0].text
            ws['R' + str(row)] = tds[1].text

        except Exception as e:
            print('Sector index error : %s' % e)
            row += 1
            continue

        row += 1

        time.sleep(DELAY)


if __name__ == '__main__':
    wb = load_fs_workbook()
    ws = wb[SP500SHEET]

    build_sp500_list()
    get_other_info()
    wb.save('sp500.xlsx')
