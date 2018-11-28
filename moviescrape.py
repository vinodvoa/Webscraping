from bs4 import BeautifulSoup
from openpyxl import Workbook

import requests
import time

# new workbook and sheet
wb = Workbook()
ws = wb.active

# initialization
row = 1
col = 1
links = []

# loop through 10 page of the website
for pages in range(1, 11):
    URL = 'https://ffmovies.ru/top-imdb' + '?page=' + str(pages)
    # print(URL)

    page = requests.get(URL)

    soup = BeautifulSoup(page.content, 'lxml')

    # get all containers housing the movies
    divtag = soup.findAll('div', class_='col-lg-3 col-md-4 col-sm-6 col-xs1-8 col-xs-12')

    # loop through each movie container to pull details
    for div in divtag:
        movie = div.find('a', class_='name')

        # movie title
        link = movie.get('href')
        ws.cell(row, col).value = movie.string
        col += 1

        # movie link
        ws.cell(row, col).value = 'https://ffmovies.ru' + link
        movielink = 'https://ffmovies.ru' + link
        links.append(movielink)
        col += 1
        # print(movie.string, link)

        # movie quality
        quality = div.find('div', class_='quality')

        if quality is not None:
            ws.cell(row, col).value = quality.text
            col += 1
            # print(quality.text)

        # Episode
        episode = div.find('div', class_='status')

        if episode is not None:
            ws.cell(row, col).value = episode.text
            # print(episode.text)

        row += 1
        col = 1

# print(links)

# get the movie details
row = 1
col = 5

# loop through movie links to get additional info
for movie in links:
    print(movie)
    page = requests.get(movie)
    soup = BeautifulSoup(page.content, 'lxml')

    # IMDb rating
    imdbtag = soup.find('div', class_='info col-md-19')

    imdb = imdbtag.div.div.span.text
    ws.cell(row, col).value = imdb
    col += 1
    # print(imdb)

    # genre
    genretag = soup.find('dl', class_='meta col-sm-12')

    genre = genretag.dd.a.text
    ws.cell(row, col).value = genre
    # print(genre)

    # releasetag = soup.find('dd', class_='rating')
    # releasetag = genretag.next_
    # print(releasetag)
    row += 1
    col = 5
    time.sleep(3)
    # break

# save workbook
wb.save('ffmoviesimdb.xlsx')
