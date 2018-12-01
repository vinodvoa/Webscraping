#!/usr/bin/env python3

from bs4 import BeautifulSoup
from openpyxl import Workbook

import requests
import time
import logging

# logger setup
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

formatter = logging.Formatter('%(levelname)s:%(name)s:%(asctime)s:%(message)s')

stream_handler = logging.StreamHandler()
stream_handler.setLevel(logging.DEBUG)
stream_handler.setFormatter(formatter)
logger.addHandler(stream_handler)

file_handler = logging.FileHandler('.log')
file_handler.setFormatter(formatter)
file_handler.setLevel(logging.DEBUG)
logger.addHandler(file_handler)

# new workbook and sheet
wb = Workbook()
ws = wb.active

# IMDb URL
URL = 'https://www.imdb.com/chart/top-english-movies?pf_rd_m=A2FGELUUNOQJNL&pf_rd_p=4da9d9a5-d299-43f2-9c53-f0efa18182cd&pf_rd_r=AYZY33SKTH1WWB3AXFEB&pf_rd_s=right-4&pf_rd_t=15506&pf_rd_i=top&ref_=chttp_ql_4'
logger.info(URL)

# get url contents
page = requests.get(URL)

# create soup
soup = BeautifulSoup(page.content, 'lxml')

# find all table rows
tr = soup.findAll('tr')
logger.debug(tr)

row = 1
col = 1
links = []

# loop through each table row and get title, movie detail link and year
for td in tr:
    titlecols = td.findAll('td', class_='titleColumn')
    logger.debug(titlecols)

    for titlecol in titlecols:
        # movie title
        title = titlecol.find('a')
        logger.info(title)
        ws.cell(row, col).value = title.text
        col += 1

        # movie detail link
        link = title.get('href')
        movielink = 'https://www.imdb.com' + link
        links.append(movielink)
        logger.debug(movielink)
        logger.debug(link)

        # movie release year
        year = titlecol.find('span')
        ws.cell(row, col).value = year.text.replace('(', '').replace(')', '')
        logger.info(year)

        row += 1
        col = 1

# loop through each table row and get movie rating
row = 1
col = 3

for td in tr:
    ratingcols = td.findAll('td', class_='ratingColumn imdbRating')
    logger.debug(ratingcols)

    # movie IMDb rating
    for ratingcol in ratingcols:
        rating = ratingcol.find('strong')
        ws.cell(row, col).value = rating.text
        logger.info(rating)

        row += 1

# loop through movie links to get pg rating, summary, duration, director and awards
row = 1
col = 4

for link in links:
    logger.debug(link)
    page = requests.get(link)
    soup = BeautifulSoup(page.content, 'lxml')

    # movie PG rating
    divtag = soup.find('div', class_='subtext')
    logger.debug(divtag)

    if len(divtag.contents[0]) > 0:
        pgrating = divtag.contents[0].strip()
    else:
        pgrating = ''

    logger.info(pgrating)
    ws.cell(row, col).value = pgrating
    col += 1

    # movie duration
    duration = divtag.time.text.strip()
    logger.info(duration)
    ws.cell(row, col).value = duration
    col += 1

    # movie summary
    divtag = soup.find('div', class_='summary_text')
    logger.debug(divtag)

    summ = divtag.text.strip()
    logger.info(summ)
    ws.cell(row, col).value = summ
    col += 1

    # movie director
    divtag = soup.find('div', class_='credit_summary_item')
    logger.debug(divtag)

    director = divtag.a.text
    logger.info(director)
    ws.cell(row, col).value = director
    col += 1

    # movie awards / nominations
    divtag = soup.find('span', class_='awards-blurb')
    logger.debug(divtag)

    spantag = str(divtag)
    logger.debug(spantag)

    if spantag.find('<b>') > 0:
        awards = divtag.b.text.strip()
    elif spantag.find('<b>') < 0:
        if divtag is not None:
            if len(divtag.contents[0]) > 0:
                awards = divtag.contents[0].strip()
        else:
            awards = ''

    logger.info(awards)
    ws.cell(row, col).value = awards

    row += 1
    col = 4
    time.sleep(1)

# save workbook
wb.save('imdbtop250.xlsx')
